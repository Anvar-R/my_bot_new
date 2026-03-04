"""Excel file parser for route data"""

import logging
from typing import List, Tuple

from openpyxl import load_workbook

logger = logging.getLogger(__name__)


async def parse_route_excel(file_path: str) -> List[Tuple[int, str, int]]:
    """
    Parse Excel file containing route data from a file path.

    Expected columns:
    - agent_id (int)
    - reg_name (str, max 50 chars)
    - visit_day (int)

    Args:
        file_path: Path to the Excel file

    Returns:
        List of tuples (agent_id, reg_name, visit_day)

    Raises:
        ValueError: If file format is invalid or required columns are missing
    """
    try:
        # Load workbook from file path
        workbook = load_workbook(file_path, data_only=True)
        sheet = workbook.active

        if sheet is None:
            raise ValueError("Excel file has no active sheet")

        # Get headers from first row
        headers = []
        for cell in sheet[1]:
            headers.append(cell.value.lower().strip() if cell.value else None)

        # Find column indices
        try:
            agent_id_idx = headers.index("agent_id") + 1
            reg_name_idx = headers.index("reg_name") + 1
            visit_day_idx = headers.index("visit_day") + 1
        except ValueError as e:
            raise ValueError(f"Missing required column. Headers found: {headers}") from e

        # Extract data rows
        routes = []
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if all(cell is None for cell in row):
                # Skip empty rows
                continue

            try:
                agent_id = int(row[agent_id_idx - 1])
                reg_name = str(row[reg_name_idx - 1]).strip()
                visit_day = int(row[visit_day_idx - 1])

                if len(reg_name) > 50:
                    raise ValueError(f"reg_name exceeds 50 characters at row {row_idx}")

                routes.append((agent_id, reg_name, visit_day))
            except (ValueError, TypeError) as e:
                logger.error(f"Error parsing row {row_idx}: {e}")
                raise ValueError(f"Invalid data at row {row_idx}: {e}") from e

        if not routes:
            raise ValueError("No data rows found in Excel file")

        logger.info(f"Successfully parsed {len(routes)} routes from Excel")
        return routes

    except Exception as e:
        logger.error(f"Error parsing Excel file: {e}")
        raise


async def parse_equip_excel(file_path: str) -> List[Tuple[int, str, str, int]]:
    """
    Parse Excel file containing equipment data from a file path.

    Expected columns:
    - agent_id (int)
    - cust_name (str, max 50 chars)
    - equip_type (str, max 20 chars)
    - visit_day (int)

    Args:
        file_path: Path to the Excel file

    Returns:
        List of tuples (agent_id, cust_name, equip_type, visit_day)

    Raises:
        ValueError: If file format is invalid or required columns are missing
    """
    try:
        # Load workbook from file path
        workbook = load_workbook(file_path, data_only=True)
        sheet = workbook.active

        if sheet is None:
            raise ValueError("Excel file has no active sheet")

        # Get headers from first row
        headers = []
        for cell in sheet[1]:
            headers.append(cell.value.lower().strip() if cell.value else None)

        # Find column indices
        try:
            agent_id_idx = headers.index("agent_id") + 1
            cust_name_idx = headers.index("cust_name") + 1
            equip_type_idx = headers.index("equip_type") + 1
            visit_day_idx = headers.index("visit_day") + 1
        except ValueError as e:
            raise ValueError(f"Missing required column. Headers found: {headers}") from e

        # Extract data rows
        equipment = []
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if all(cell is None for cell in row):
                # Skip empty rows
                continue

            try:
                agent_id = int(row[agent_id_idx - 1])
                cust_name = str(row[cust_name_idx - 1]).strip()
                equip_type = str(row[equip_type_idx - 1]).strip()
                visit_day = int(row[visit_day_idx - 1])

                if len(cust_name) > 50:
                    raise ValueError(f"cust_name exceeds 50 characters at row {row_idx}")
                if len(equip_type) > 20:
                    raise ValueError(f"equip_type exceeds 20 characters at row {row_idx}")

                equipment.append((agent_id, cust_name, equip_type, visit_day))
            except (ValueError, TypeError) as e:
                logger.error(f"Error parsing row {row_idx}: {e}")
                raise ValueError(f"Invalid data at row {row_idx}: {e}") from e

        if not equipment:
            raise ValueError("No data rows found in Excel file")

        logger.info(f"Successfully parsed {len(equipment)} equipment records from Excel")
        return equipment

    except Exception as e:
        logger.error(f"Error parsing Excel file: {e}")
        raise
